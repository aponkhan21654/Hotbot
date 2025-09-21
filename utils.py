import os
import logging
import asyncio
import aiohttp
import random
import string
import re
from datetime import datetime
from functools import wraps
from openpyxl import load_workbook, Workbook
from config import SERVICE_FILES, CODE_FORMATS, ADMIN_IDS

# Logging setup
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)
logger = logging.getLogger(__name__)

# Global data & state
user_sessions = {}
session_timeouts = {}

# Decorators
def admin_only(func):
    @wraps(func)
    async def wrapper(update, context, *args, **kwargs):
        if update.effective_user and update.effective_user.id in ADMIN_IDS:
            return await func(update, context, *args, **kwargs)
        else:
            await update.effective_message.reply_text(
                "Access Denied. You don't have permission to use this command."
            )
    return wrapper

# XLSX File Handling
def _read_excel(filename):
    """Read data from Excel file"""
    if not load_workbook or not os.path.exists(filename):
        return []
    
    try:
        wb = load_workbook(filename)
        ws = wb.active
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        return data
    except Exception as e:
        logger.error(f"Error reading Excel file {filename}: {e}")
        return []

def _write_excel(filename, data):
    """Write data to Excel file"""
    if not Workbook:
        logger.error("openpyxl not found. Cannot write XLSX.")
        return False
    
    try:
        wb = Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        wb.save(filename)
        return True
    except Exception as e:
        logger.error(f"Error writing Excel file {filename}: {e}")
        return False

def get_stock_count(service_key):
    """Get stock count for a service"""
    filename = SERVICE_FILES.get(service_key)
    if not filename:
        return 0
    data = _read_excel(filename)
    return max(0, len(data) - 1)  # Subtract header row

def get_rows_for_purchase(service_key, count):
    """Get rows for purchase from Excel file"""
    filename = SERVICE_FILES.get(service_key)
    if not filename:
        return None
    
    data = _read_excel(filename)
    if len(data) <= 1:  # Only header or empty
        return None
    
    # Return requested count of rows (excluding header)
    return data[1:count+1] if count <= len(data)-1 else data[1:]

def remove_purchased_rows(service_key, count):
    """Remove purchased rows from Excel file"""
    filename = SERVICE_FILES.get(service_key)
    if not filename:
        return False
    
    data = _read_excel(filename)
    if len(data) <= 1:  # Only header or empty
        return False
    
    # Keep header and rows after the purchased ones
    new_data = [data[0]] + data[count+1:]
    return _write_excel(filename, new_data)

def append_excel_data(service_key, new_rows_with_header):
    """Append data to Excel file"""
    filename = SERVICE_FILES.get(service_key)
    if not filename:
        return False
    
    # If file doesn't exist, create with the provided data
    if not os.path.exists(filename):
        return _write_excel(filename, new_rows_with_header)
    
    # Read existing data and append new rows
    existing_data = _read_excel(filename)
    # Skip header from new data if existing file already has one
    if existing_data and new_rows_with_header:
        new_data = existing_data + new_rows_with_header[1:]
    else:
        new_data = existing_data + new_rows_with_header
    
    return _write_excel(filename, new_data)

def create_user_download_file(rows_data, service_key):
    """Create a download file for user"""
    import tempfile
    import os
    
    # Create a temporary file
    fd, temp_path = tempfile.mkstemp(suffix='.txt')
    
    try:
        with os.fdopen(fd, 'w') as tmp:
            for row in rows_data:
                line = '|'.join(str(cell) for cell in row)
                tmp.write(line + '\n')
        
        return temp_path
    except Exception as e:
        logger.error(f"Error creating download file: {e}")
        return None

def calculate_discount(quantity):
    """Calculate discount based on quantity"""
    from database import get_discount_settings
    
    discounts = get_discount_settings()
    applicable_discount = 0
    
    for min_qty, discount_percent in sorted(discounts, key=lambda x: x[0], reverse=True):
        if quantity >= min_qty:
            applicable_discount = discount_percent
            break
    
    return applicable_discount

# API Functions
async def fetch_code_from_api(api_url, params):
    """Fetch code from API"""
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(api_url, params=params, timeout=30) as response:
                response.raise_for_status()
                data = await response.json()
                logger.info(f'API Response from {api_url}: {data}')
                return data.get('code'), data
    except aiohttp.ClientError as ce:
        logger.error(f'API connection error for {api_url}: {ce}')
        return None, {'status': 'error', 'message': f'API connection error: {ce}'}
    except asyncio.TimeoutError:
        logger.error(f'API request timed out for {api_url}')
        return None, {'status': 'error', 'message': 'API request timed out.'}
    except Exception as e:
        logger.error(f'Unexpected API error for {api_url}: {e}')
        return None, {'status': 'error', 'message': f'Unexpected API error: {e}'}

# Session Management
def clear_user_session(user_id):
    """Clear user session"""
    if user_id in user_sessions:
        del user_sessions[user_id]

async def set_session_timeout(user_id, context, minutes=15):
    """Set session timeout"""
    clear_user_session(user_id)
    
    # Cancel any existing timeout for this user
    if user_id in session_timeouts:
        session_timeouts[user_id].cancel()
    
    # Set new timeout
    async def timeout_callback():
        if user_id in user_sessions:
            del user_sessions[user_id]
            try:
                await context.bot.send_message(
                    chat_id=user_id, 
                    text=f"Session expired after {minutes} minutes of inactivity. Please start again."
                )
            except Exception as e:
                logger.error(f"Error sending timeout message: {e}")
    
    # Schedule the timeout
    task = asyncio.create_task(asyncio.sleep(minutes * 60))
    session_timeouts[user_id] = task
    task.add_done_callback(lambda _: timeout_callback())

# Referral Functions
def generate_referral_code(user_id):
    """Generate a referral code"""
    return f"REF{user_id}{random.randint(1000, 9999)}"

def get_or_create_referral_code(user_id):
    """Get or create a referral code for user"""
    from database import get_user_data, db_execute
    
    user_data = get_user_data(user_id)
    if user_data and user_data[3]:  # referral_code field
        return user_data[3]
    else:
        referral_code = generate_referral_code(user_id)
        db_execute('UPDATE users SET referral_code = ? WHERE user_id = ?', (referral_code, user_id))
        return referral_code

def get_referral_link(user_id, bot_username="your_bot_username"):
    """Get referral link"""
    referral_code = get_or_create_referral_code(user_id)
    return f"https://t.me/{bot_username}?start={referral_code}"

def handle_referral_signup(user_id, referral_code):
    """Handle referral signup"""
    from database import db_execute
    
    referrer_info = db_execute('SELECT user_id FROM users WHERE referral_code = ?', (referral_code,), fetchone=True)
    if referrer_info:
        referrer_id = referrer_info[0]
        db_execute('UPDATE users SET total_referrals = total_referrals + 1 WHERE user_id = ?', (referrer_id,))
        db_execute('UPDATE users SET referred_by = ? WHERE user_id = ?', (referrer_id, user_id))

def get_referral_stats(user_id):
    """Get referral stats"""
    from database import db_execute, get_referral_settings
    
    total_refs = db_execute('SELECT total_referrals FROM users WHERE user_id = ?', (user_id,), fetchone=True)
    total_earnings = db_execute('SELECT SUM(reward_amount) FROM referral_rewards WHERE referrer_id = ? AND status = "approved"', (user_id,), fetchone=True)
    pending_rewards = db_execute('SELECT COUNT(*) FROM referral_rewards WHERE referrer_id = ? AND status = "pending"', (user_id,), fetchone=True)
    referrer_bonus, referred_bonus = get_referral_settings()
    
    return {
        'total_refs': total_refs[0] if total_refs else 0,
        'total_earnings': total_earnings[0] if total_earnings else 0.0,
        'pending_rewards': pending_rewards[0] if pending_rewards else 0,
        'referrer_bonus': referrer_bonus,
        'referred_bonus': referred_bonus
    }